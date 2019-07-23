from tkinter import *
import datetime
import openpyxl
from openpyxl import load_workbook

class SlowCH_Manager(Canvas):
    """ 
    Manages a variable number of slow channel massages
    """
    def __init__(self,master=None,**kwargs):
        Canvas.__init__(self,master,**kwargs,width = 1000,height = 550)
        self.frame = Frame(self)
        self.create_window(0,0,anchor=N+W,window=self.frame)
        self.row = 0
        self.widgets = []
        self.max = 100
        self._init_entries()

    def _init_entries(self):
        """
        initialize the input area with labels and perhaps default values
        """
        label_id  = Label(self.frame, text='Name').grid(row = self.row, column = 1)
        label_check  = Label(self.frame, text='MI?').grid(row = self.row, column = 2)
        label_msg = Label(self.frame, text='Quantity').grid(row = self.row, column = 3)
        label_pri = Label(self.frame, text='Price').grid(row = self.row, column = 4)
        label_buy = Label(self.frame, text='Register').grid(row = self.row, column = 5)
        self.row += 1


    def add_entry(self):
        """
        Dynamically add entry to GUI until max number of entries is arrived.
        By SENT specification max 100 slow channel messages are allowed.
        """
        if len(self.widgets) >= self.max:
            print('Im full')
        else:
            label = Label(self.frame, text=str(len(self.widgets))).grid(row = self.row, column = 0)
            entry_id = Entry(self.frame)
            entry_id.grid(row = self.row, column = 1)
            entry_data = Entry(self.frame)
            var = IntVar()
            chk = Checkbutton(self.frame, variable=var)
            chk.grid(row = self.row, column = 2)
            entry_data.grid(row = self.row, column =3)
            entry_pri = Entry(self.frame)
            entry_pri.grid(row = self.row, column = 4)
            entry_buy = Entry(self.frame)
            entry_buy.grid(row = self.row, column = 5)
            self.row += 1
            self.widgets.append(entry_id)
            all_entries.append( (entry_id, var, entry_data, entry_pri, entry_buy) )


    def _ypos(self):
        return sum(x.winfo_reqheight() for x in self.widgets)


if __name__ == "__main__":
    root = Tk()
    w = 1300 # width for the Tk root
    h = 700 # height for the Tk root

    #get screen width and height
    ws = root.winfo_screenwidth() # width of the screen
    hs = root.winfo_screenheight() # height of the screen

    # calculate x and y coordinates for the Tk root window
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)

    # set the dimensions of the screen 
    # and where it is placed
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))

    manager = SlowCH_Manager(root)
    manager.grid(row=0,column=0)

    scroll = Scrollbar(root)
    scroll.grid(row=0,column=5,sticky=N+S)

    manager.config(yscrollcommand = scroll.set)
    scroll.config(command=manager.yview)
    manager.configure(scrollregion = manager.bbox("all"))

    def command():
        manager.add_entry()
        # update scrollregion
        a = manager.bbox("all")
        b = (0,0,30,30)
        big_region = tuple(map(sum,zip(a,b)))
        manager.configure(scrollregion = big_region)
        
    def showEntries():

        rows = len(all_entries)

        today_master = str(datetime.datetime.now().date()) + ".xlsx"

        #master log, not editable
        wb = openpyxl.Workbook(today_master)
        wb.save(today_master)
        wb = openpyxl.load_workbook(today_master)
        wb.create_sheet('Stock')
        wb.save(today_master)
        ws = wb.active
        wb.get_sheet_names()
        sheet = wb.get_sheet_by_name('Sheet')
        stock = wb.get_sheet_by_name('Stock')

       
        for number, (entry_id, var, entry_data, entry_pri,entry_buy) in enumerate(all_entries):
            sheet['A' + str(number+1)].value = entry_id.get()
            sheet['B' + str(number+1)].value = entry_data.get()
            sheet['C' + str(number+1)].value = entry_pri.get()
            stock['A' + str(number+1)].value = entry_id.get()
            stock['B' + str(number+1)].value = entry_data.get()
            stock['C' + str(number+1)].value = entry_pri.get()
            wb.save(today_master)
            
    def confirm_price():
        item_total = 0
        MI_total = 0
        for number, (entry_id, var, entry_data, entry_pri, entry_buy) in enumerate(all_entries):
            a = eval(entry_pri.get())*eval(entry_buy.get())
            item_total = item_total + a
            if var.get() == 1:
                b = eval(entry_pri.get())*eval(entry_buy.get())
                MI_total += b
            today_master = str(datetime.datetime.now().date()) + ".xlsx"
            wb = openpyxl.load_workbook(today_master)
            stock = wb.get_sheet_by_name('Stock')
            stock['B' + str(number+1)].value = float(stock['B' + str(number+1)].value) - float(entry_buy.get())
            wb.save(today_master)
            
    def inventory():
        rows = len(all_entries)
        top = Toplevel()
        ksbar=Scrollbar(top, orient=VERTICAL)
        ksbar.grid(row=0, column=1, sticky="ns")
        popCanv = Canvas(top, width=600, height = 800,
        scrollregion=(0,0,500,800)) #width=1256, height = 1674)
        popCanv.grid(row=0, column=0, sticky="nsew") #added sticky
        ksbar.config(command=popCanv.yview)
        popCanv.config(yscrollcommand = ksbar.set)
        today_master = str(datetime.datetime.now().date()) + ".xlsx"
        wb = openpyxl.load_workbook(today_master)
        stock = wb.get_sheet_by_name('Stock')
        for i in range(1,rows+1):
            Label(popCanv, text= 'Product').grid(row = 0, column = 0)
            Label(popCanv, text= str(stock['A' + str(i)].value)).grid(row = str(i), column = 0)
            Label(popCanv, text= 'Remaining Quantity').grid(row = 0, column = 1)
            Label(popCanv, text= str(stock['B' + str(i)].value)).grid(row = str(i), column = 1)

    def revenue():
        rows = len(all_entries)
        top = Toplevel()
        ksbar=Scrollbar(top, orient=VERTICAL)
        ksbar.grid(row=0, column=1, sticky="ns")
        popCanv = Canvas(top, width=600, height = 800,
        scrollregion=(0,0,500,800)) #width=1256, height = 1674)
        popCanv.grid(row=0, column=0, sticky="nsew") #added sticky
        ksbar.config(command=popCanv.yview)
        popCanv.config(yscrollcommand = ksbar.set)
        today_master = str(datetime.datetime.now().date()) + ".xlsx"
        wb = openpyxl.load_workbook(today_master)
        stock = wb.get_sheet_by_name('Stock')
        master = wb.get_sheet_by_name('Sheet')
        total = 0
        for i in range(1,rows+1):
            Label(popCanv, text= 'Product').grid(row = 0, column = 0)
            Label(popCanv, text= str(stock['A' + str(i)].value)).grid(row = str(i), column = 0)
            Label(popCanv, text= 'TOTAL').grid(row = str(rows+2), column = 0)
            Label(popCanv, text= 'Revenue').grid(row = 0, column = 1)
            per_item = (float(master['B' + str(i)].value)-float(stock['B' + str(i)].value))*float(master['C' + str(i)].value)
            total += per_item
            Label(popCanv, text= '$'+str(per_item)).grid(row = str(i), column = 1)
            Label(popCanv, text= '$'+str(total)).grid(row = str(rows+2), column = 1)


    def calculate():
        top = Toplevel()
        item_total = 0
        MI_total = 0
        for number, (entry_id, var, entry_data, entry_pri, entry_buy) in enumerate(all_entries):
            a = eval(entry_pri.get())*eval(entry_buy.get())
            item_total = item_total + a
            if var.get() == 1:
                b = eval(entry_pri.get())*eval(entry_buy.get())
                MI_total += b
        Label(top, text="Totals").grid(row = 0, column = 1)                 
        Label(top, text="MI$"+str(MI_total)).grid(row = 1, column = 1)                 
        Label(top, text="$"+str(item_total)).grid(row = 1, column = 2)
        b = Button(top, text = "Confirm sale", command = confirm_price)
        b.grid(row=1,column=0)
        
                
    all_entries = []
    b = Button(root, text = "Add Product", command = command)
    b.grid(row=1,column=0)
    c = Button(root, text = "Submit Inventory", command = showEntries)
    c.grid(row=2,column=0)
    d = Button(root, text = "Calculate Sale", command = calculate)
    d.grid(row=3,column=0)
    e = Button(root, text = "Check Inventory", command = inventory)
    e.grid(row = 1, column = 1)
    f = Button(root, text = "Check Revenue", command = revenue)
    f.grid(row=2,column=1)

    root.mainloop()
